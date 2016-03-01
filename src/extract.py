from bs4 import BeautifulSoup
import urllib2,sys
urls = ["http://cloud-computing.softwareinsider.com/l/265/Atlantic-Net","http://cloud-computing.softwareinsider.com/l/162/CloudSigma-Holding-AG","http://cloud-computing.softwareinsider.com/l/18/Engine-Yard-Inc","http://cloud-computing.softwareinsider.com/l/17/Microsoft-Corporation","http://cloud-computing.softwareinsider.com/l/143/WorkXpress","http://cloud-computing.softwareinsider.com/l/150/Lunacloud-Lda","http://cloud-computing.softwareinsider.com/l/5/Amazon-Inc","http://cloud-computing.softwareinsider.com/l/14/Rackspace-US-Inc","http://cloud-computing.softwareinsider.com/l/96/Dimension-Data","http://cloud-computing.softwareinsider.com/l/313/Digia-Plc","http://cloud-computing.softwareinsider.com/l/181/iWeb-Technologies-Inc","http://cloud-computing.softwareinsider.com/l/262/FireHost-Inc","http://cloud-computing.softwareinsider.com/l/307/Devart-S-R-O","http://cloud-computing.softwareinsider.com/l/129/TheCompuLab-Corp","http://cloud-computing.softwareinsider.com/l/146/VPS-NET","http://cloud-computing.softwareinsider.com/l/183/BlueiTech-Global-Internet-Solutions","http://cloud-computing.softwareinsider.com/l/230/AppScale-Systems-Inc","http://cloud-computing.softwareinsider.com/l/292/DigitalOcean-Inc","http://cloud-computing.softwareinsider.com/l/118/Progress-Software-Corporation","http://cloud-computing.softwareinsider.com/l/300/Google-Inc","http://cloud-computing.softwareinsider.com/l/256/Tsuru","http://cloud-computing.softwareinsider.com/l/2/GoGrid-LLC","http://cloud-computing.softwareinsider.com/l/258/Voxoz","http://cloud-computing.softwareinsider.com/l/259/Tabcou-LLC","http://cloud-computing.softwareinsider.com/l/260/The-Apache-Software-Foundation","http://cloud-computing.softwareinsider.com/l/261/Selectel-Ltd","http://cloud-computing.softwareinsider.com/l/6/Joyent-Inc","http://cloud-computing.softwareinsider.com/l/263/AppHarbor-Inc","http://cloud-computing.softwareinsider.com/l/264/Brightbox-Systems-Ltd","http://cloud-computing.softwareinsider.com/l/9/Flexiant-Limited","http://cloud-computing.softwareinsider.com/l/266/ArubaCloud","http://cloud-computing.softwareinsider.com/l/267/Belgacom","http://cloud-computing.softwareinsider.com/l/268/CollabNet-Inc","http://cloud-computing.softwareinsider.com/l/269/CloudFlare-Inc","http://cloud-computing.softwareinsider.com/l/270/Nimbix-Inc","http://cloud-computing.softwareinsider.com/l/15/Google-App-Engine","http://cloud-computing.softwareinsider.com/l/271/Serversaurus","http://cloud-computing.softwareinsider.com/l/16/OneNeck-IT-Solutions-LLC","http://cloud-computing.softwareinsider.com/l/272/GMO-CLOUD-AMERICA-INC","http://cloud-computing.softwareinsider.com/l/273/Pivotal-Software-Inc","http://cloud-computing.softwareinsider.com/l/274/Cloudant","http://cloud-computing.softwareinsider.com/l/19/Salesforce-com-inc","http://cloud-computing.softwareinsider.com/l/275/Acquia-Inc","http://cloud-computing.softwareinsider.com/l/276/Amazon-Web-Services-Inc","http://cloud-computing.softwareinsider.com/l/277/Clever-Cloud-SAS","http://cloud-computing.softwareinsider.com/l/22/Verizon-Terremark","http://cloud-computing.softwareinsider.com/l/278/CloudBees-Inc","http://cloud-computing.softwareinsider.com/l/24/IBM","http://cloud-computing.softwareinsider.com/l/280/Eldarion-Inc","http://cloud-computing.softwareinsider.com/l/25/AT-And-T-Inc","http://cloud-computing.softwareinsider.com/l/281/Heroku-Inc","http://cloud-computing.softwareinsider.com/l/282/Hewlett-Packard-Development-Company-L-P","http://cloud-computing.softwareinsider.com/l/283/Modulus","http://cloud-computing.softwareinsider.com/l/284/Anchora-Inc","http://cloud-computing.softwareinsider.com/l/285/SoftLayer-Technologies-Inc","http://cloud-computing.softwareinsider.com/l/286/Red-Hat-Inc","http://cloud-computing.softwareinsider.com/l/287/Pagoda-Box","http://cloud-computing.softwareinsider.com/l/288/Pogoapp","http://cloud-computing.softwareinsider.com/l/33/Boomi-Inc","http://cloud-computing.softwareinsider.com/l/289/Snapp-Systems"]
'''for url in urls:
    try:
        req = urllib2.Request(url, headers={'User-Agent' : 'Magic Browser'})
        con = urllib2.urlopen(req)
        page = con.read()
        soup = BeautifulSoup(page)
        section = soup.find('section',{'data-section-id':'2'})
        headings = section.findAll('div',{'class':'dd-overview-label'})
        values = section.findAll('noscript')
        print headings,values
    except:
        print url'''

url = urls[0]
try:
    req = urllib2.Request(url, headers={'User-Agent' : 'Magic Browser'})
    con = urllib2.urlopen(req)
    page = con.read()
    soup = BeautifulSoup(page)
    section = soup.find('section',{'data-section-id':'2'})
    headings = section.findAll('div',{'class':'dd-overview-label'})
    values = section.findAll('noscript')
    print headings,values
except:
    print url
from openpyxl import load_workbook
wb = load_workbook(filename = 'TabulatedSLA.xlsx')
ws = wb['Sheet1']
sname = url.split('/')[-1]
print sname
for r in range(len(ws.rows)):
    if(ws.cell(row=r,column=2).value == sname):
        for col in range(len(ws.rows[r])):
            if(ws.cell(row=0,column=col).value == headings[0].text):
                print "changing"
                ws.cell(row=r,column=col).value = values[0].text
            if(ws.cell(row=0,column=col).value == headings[1].text):
                print "changing"
                ws.cell(row=r,column=col).value = values[1].text
            if(ws.cell(row=0,column=col).value == headings[2].text):
                print "changing"
                ws.cell(row=r,column=col).value = values[2].text
            if(ws.cell(row=0,column=col).value == headings[3].text):
                print "changing"
                ws.cell(row=r,column=col).value = values[3].text
wb.save('TabulatedSLA.xlsx')
print "done"
